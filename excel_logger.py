import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

FILE_NAME = "bot_logs.xlsx"

def _get_workbook():
    """Открывает существующий или создает новый файл Excel"""
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        
        # Настраиваем лист Пользователей
        ws_users = wb.active
        ws_users.title = "Users"
        ws_users.append(["Дата", "Время", "IT-код", "Имя Фамилия", "Действие", "Товар", "Кол-во", "Комментарий", "Статус"])
        
        # Настраиваем лист Админов
        ws_admin = wb.create_sheet("Admins")
        ws_admin.append(["Дата", "Время", "Admin ID", "Действие", "Детали"])
        
        wb.save(FILE_NAME)
        return wb
    else:
        return load_workbook(FILE_NAME)

def log_user_action(user_obj, action, item_name="-", qty="-", comment="-", status="-"):
    """Запись действий пользователя"""
    try:
        wb = _get_workbook()
        ws = wb["Users"]
        
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        full_name = f"{user_obj.first_name} {user_obj.last_name or ''}".strip()
        
        ws.append([
            date_str, 
            time_str, 
            user_obj.it_code, 
            full_name, 
            action, 
            item_name, 
            qty, 
            comment, 
            status
        ])
        
        wb.save(FILE_NAME)
    except Exception as e:
        print(f"❌ Ошибка записи лога (User): {e}")

def log_admin_action(admin_id, action, details):
    """Запись действий админа"""
    try:
        wb = _get_workbook()
        # Если вдруг листа нет (файл старый), создадим
        if "Admins" not in wb.sheetnames:
            ws = wb.create_sheet("Admins")
            ws.append(["Дата", "Время", "Admin ID", "Действие", "Детали"])
        else:
            ws = wb["Admins"]
        
        now = datetime.now()
        date_str = now.strftime("%Y-%m-%d")
        time_str = now.strftime("%H:%M:%S")
        
        ws.append([
            date_str,
            time_str,
            str(admin_id),
            action,
            details
        ])
        
        wb.save(FILE_NAME)
    except Exception as e:
        print(f"❌ Ошибка записи лога (Admin): {e}")